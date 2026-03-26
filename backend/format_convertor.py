import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import *
import datetime as dt


def _normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _pick_existing_column(df, candidates):
    for col in candidates:
        if col in df.columns:
            return col
    return None


def type_convert(args):
    if str(args) != "nan":
        try:
            return '%d' % int(args)
        except ValueError:
            return str(args)
    else:
        return ""


def format_convertor(input_filename, read_path, save_path):
    '''
    Input: a list of filenames

    '''
    ID = []
    pat_ID = []
    hosp_ID = []
    short_ID = []
    gender = []
    age = []
    trio = []
    index = []
    doctor = []
    hospital = []
    name = []
    birth = []
    note = []
    tel1 = []
    tel2 = []
    null = []

    x = len(input_filename)
    for file_name in input_filename:
        # read dataframe
        file_type = file_name.split('.')[-1:][0]

        if file_type == "csv":
            df_old = pd.read_csv(read_path + '/' +file_name, dtype=str)
        elif file_type == "xlsx":
            df_old = pd.read_excel(read_path + '/' + file_name, dtype=str)
        elif file_type == "tsv" or file_type == "txt":
            df_old = pd.read_table(read_path + '/' + file_name, dtype=str)
        else:
            print("文件类型有误，仅支持csv,xlsx,txt格式文件")

        # 统一处理表头，避免 BOM/空格导致匹配失败
        df_old.columns = [str(col).replace("\ufeff", "").strip() for col in df_old.columns]

        required_columns = [
            "姓名", "P-SID", "身份证号", "首诊医生", "就诊医院",
            "就诊卡号", "病案号", "联系电话", "备用联系电话", "其他", "性别"
        ]
        missing_columns = [col for col in required_columns if col not in df_old.columns]
        if missing_columns:
            raise KeyError(f"缺少必要列: {', '.join(missing_columns)}")

        trio_col = _pick_existing_column(df_old, ["家系组", "家系角色", "组织"])
        if trio_col is None:
            raise KeyError("缺少家系信息列: 家系组/家系角色/组织")

        # 筛选
        valid_name = df_old["姓名"].fillna("").astype(str).str.strip().str.len() > 0
        df_input = df_old[valid_name]
        df_input.reset_index(drop=True, inplace=True)


        for i in range(len(df_input)):
            short_ID.append(df_input["P-SID"][i])
            name.append(_normalize_text(df_input["姓名"][i]))
            ID.append(type_convert(df_input["身份证号"][i]))
            doctor.append(_normalize_text(df_input["首诊医生"][i]))
            hospital.append(_normalize_text(df_input["就诊医院"][i]))
            birth.append(str(df_input["身份证号"][i])[6:14])
            pat_ID.append(type_convert(df_input["就诊卡号"][i]))
            hosp_ID.append(type_convert(df_input["病案号"][i]))
            tel1.append(type_convert(df_input["联系电话"][i]))
            tel2.append(type_convert(df_input["备用联系电话"][i]))
            note.append(_normalize_text(df_input["其他"][i]))
            null.append("")

            i_gender = _normalize_text(df_input["性别"][i])
            if i_gender == "女":
                gender.append("女")
            elif i_gender == "男":
                gender.append("男")
            else:
                gender.append("未知")

            trio_i = _normalize_text(df_input[trio_col][i])
            if trio_i in ["", "无家系", "无", "否", "散发", "单例"]:
                trio.append("否")
            else:
                trio.append("是")

            b = str(df_input["身份证号"][i])[6:14]  # 出生日期
            a = pd.to_datetime(b)
            i_age = dt.date.today().year - a.year
            age.append(i_age)

    # 修复index
    for e in range(len(name)):
        e += 1
        index.append(e)

    df_o1 = pd.DataFrame(
        {"病案号": hosp_ID, "就诊医生": doctor, "*受检人姓名": name, "0": null, "*性别": gender, "身份证号": ID, "*出生日期": birth,
         "年龄": age,
         "病人ID": pat_ID, "*取样日期": null, "00": null, "样本来源": null, "备注": note, "联系电话1": tel1, "联系电话2": tel2})
    df_o2 = pd.DataFrame(
        {"病案号": hosp_ID, "就诊医生": doctor, "*受检人姓名": name, "0": null, "*性别": gender, "身份证号": ID, "*出生日期": birth,
         "年龄": age,
         "病人ID": pat_ID, "*取样日期": null, "*取样量": null, "取样来源": null, "*单位": hospital, "*单位描述": null, "000": null,
         "是否合格": null, "0000": null, "*是否家系": trio, "*外部样本号": null, "00000": null, "*是否散发": null, "000000": null,
         "处理结果": null, "处理意见": null, "备注": note, "联系电话1": tel1, "联系电话2": tel2})
    df_o3 = pd.DataFrame(
        {"序号": index, "*样本源": name, "项目内ID": short_ID, "患者ID": pat_ID, "*样本源来源": null, "所属机构": hospital,
         "病案号": hosp_ID,
         "性别": gender, "出生日期": birth, "研究分组": null})
    col1 = ["病案号", "就诊医生", "*受检人姓名", "0", "*性别", "身份证号", "*出生日期", "年龄", "病人ID", "*取样日期", "0", "样本来源", "备注", "联系电话1",
            "联系电话2"]
    col2 = ["病案号", "就诊医生", "*受检人姓名", "0", "*性别", "身份证号", "*出生日期", "年龄", "病人ID", "*取样日期", "*取样量", "取样来源", "*单位",
            "*单位描述",
            "0", "是否合格", "0", "*是否家系", "*外部样本号", "0", "*是否散发", "0", "*处理结果", "处理意见", "备注", "联系电话1", "联系电话2"]
    df_o1.columns = col1
    df_o2.columns = col2

    df_o1.to_excel(save_path + "/样本登记_" + str(dt.date.today()) + ".xlsx", index=False)
    df_o2.to_excel(save_path + "/样本接收_" + str(dt.date.today()) + ".xlsx", index=False)
    df_o3.to_excel(save_path + "/捐献者_" + str(dt.date.today()) + ".xlsx", startrow=1, index=False)


    # 组装'捐献者'绝对路径
    file = save_path + "/捐献者_" + str(dt.date.today()) + ".xlsx"
    wb = load_workbook(file)
    ws = wb["Sheet1"]
    ws.merge_cells("A1:J1")
    ws["A1"] = "填表说明：\n1、带红' * '字段为必填项。\n2、院内捐献者患者ID必需填写，所属机构为当前医院，模板中设置无效。院外捐献者所属机构必需填写。\n3、下拉列表只能进行选择，填写无效。\n4、出生日期格式为：2021-01-01。"
    ws["A1"].font = Font('宋体', size=12, bold=True, italic=False, strike=False, color='000000')
    ws["A1"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws["A1"].fill = PatternFill('solid', fgColor="FFFF00")
    ws.row_dimensions[1].height = 100
    ws.column_dimensions['A'].width = 6.44
    ws.column_dimensions['B'].width = 13.44
    ws.column_dimensions['C'].width = 13.44
    ws.column_dimensions['D'].width = 8.44
    ws.column_dimensions['E'].width = 13.44
    ws.column_dimensions['F'].width = 13.44
    ws.column_dimensions['G'].width = 7.44
    ws.column_dimensions['H'].width = 4.44
    ws.column_dimensions['I'].width = 10.44
    ws.column_dimensions['J'].width = 10.44
    k = 0
    for cell in ws[2]:
        cell.fill = PatternFill('solid', fgColor="87CEEB")
    wb.save(file)
